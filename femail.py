import email
import imaplib
import os
import mimetypes
import bs4
from bs4 import BeautifulSoup
import types
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import quopri
import datetime
from datetime import datetime
import urllib.request
import urllib.parse

#initializing username and password for the account
username = 'input email address'
password = 'input password'
mail = imaplib.IMAP4_SSL('outlook.office365.com')
mail.login(username, password)

#mailboxes contain the folders present in the users account
result, mailboxes = mail.list()
#select the particular folder you want
mail.select("Hello")

#the emails in the folder selected are assigned unique ids
#You can specify the subject by replacing "ALL" https://yuji.wordpress.com/2011/06/22/python-imaplib-imap-example-with-gmail/
result, data = mail.uid('search', None, "ALL")

#the uids are split into a list of strings
inbox_item_list = data[0].split()

#iterating over emails present in the folder to get required info about the email
for item in inbox_item_list:
	#get all the information of the email in raw format
	result2, email_data = mail.uid('fetch', item, '(RFC822)')
	raw_email = email_data[0][1].decode("utf-8")
	email_message = email.message_from_string(raw_email)
	from_ = email_message['From']
	subject_ = email_message['Subject']
	date_ = email_message['date']
	counter = 1
for part in email_message.walk():
	if part.get_content_maintype() == "multipart":
		continue
	filename = part.get_filename()
	if not filename:
		ext = mimetypes.guess_extension(part.get_content_type())
	if not ext:
		ext ='.bin'
		filename = 'msg-part-%08d%s' %(counter,ext)
	counter +=1
	content_type = part.get_content_type()
	if "html" in content_type:
		#get html part of the email
		html_ = part.get_payload()
		#remove the 3d decoded elements that come with the email because of the way they are processed when travelling
		decoded_string = quopri.decodestring(html_)
		#to read an html tabel. read_html gives a list of dataframes
		df_list = pd.read_html(decoded_string, header = 0)
		#get the desired table by accessing the table using the index. this will give you a pandasdataframe
		df = df_list[0]
		#remove the rows that have characters
		df = df[~df['Sr'].str.contains("[a-zA-Z]").fillna(False)]
		#reset index
		df = df.reset_index(drop=True)
		#to check the dataframe : 
		#df.to_excel("final.xlsx", sheet_name = 'Sheet_name_1')

#use the openpyxl library to write to an existing excel sheet by making its object
wb = openpyxl.load_workbook('final.xlsx')
#sheet = wb.get_sheet_by_name('Sheet1')
sheet = wb["Sheet1"]

#get max rows in the excel file
max = sheet.max_row

#make two lists to compare the req no between the two
test_list = list(df['Req No'])
test_list2 = []

for rowNum in range(2, max+1):
		test_list2.append((sheet.cell(row=rowNum, column = 2).value).encode('ascii', 'ignore'))

#create a list to get the values that are not in the excel file
main_list = np.setdiff1d(test_list,test_list2)
et =[]

#add those rows that are not in the excel file
for i in main_list:
	et = ((df.iloc[df.index[df['Req No'] == i].tolist(),:6]).values.tolist())
	sheet.append(et[0])
#format the first column representing Serial no. 
p = 1
for rowNum in range(2,sheet.max_row+1):
	sheet.cell(row= rowNum, column =1).value = p
	p = p + 1

#get the dates from the remarks to compare with the dates in excel file
a = df.columns.values[6]
b = df.columns.values[7]
dt = (a[-4:])
dt2 = (b[-4:])

#format the dates so as to match the dates in the excel file
c_year = str(datetime.now().year)
dt = dt + '/' +c_year
dt2 = dt2 + '/' +c_year

#add the comments to the appropriate cells
for s, rowt in df.iterrows():
	for rowNum in range(1,sheet.max_row+1):
		if rowt["Req No"] == sheet.cell(row=rowNum, column =2).value:
			for colNum in range(1, sheet.max_column+1):
				if type(sheet.cell(row=1, column= colNum).value) == datetime:
					if dt in (sheet.cell(row=1, column= colNum).value).strftime("%m/%d/%Y"):
						sheet.cell(row=rowNum, column= colNum).value = rowt[a]

#add the comments to the appropriate cells
for s, rowt in df.iterrows():
	for rowNum in range(1,sheet.max_row+1):
		if rowt["Req No"] == sheet.cell(row=rowNum, column =2).value:
			for colNum in range(1, sheet.max_column+1):
				if type(sheet.cell(row=1, column= colNum).value) == datetime:
					if dt2 in (sheet.cell(row=1, column= colNum).value).strftime("%m/%d/%Y"):
						sheet.cell(row=rowNum, column= colNum).value = rowt[b]

#save the chnages made to the excel workbook
wb.save('final.xlsx')







